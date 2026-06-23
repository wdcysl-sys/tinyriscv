"""
Excel 读取模块

从 power_domain_iso.xlsx 解析 ISO 信号信息，包括:
- 根据 column_map.json 映射列名
- Forward-fill 处理合并单元格（空单元格 = 继承上一行）
- 输出 List[IsoSignal]
"""

from typing import Dict, List

from .models import IsoSignal


def parse_excel(filepath: str, column_map: Dict[str, str] = None) -> List[IsoSignal]:
    """
    读取 Excel 文件，解析出 ISO 信号列表

    参数:
      filepath: Excel 文件路径
      column_map: 列名映射 {内部字段名: Excel列名}，如果为 None 则不映射

    返回:
      List[IsoSignal] — 每个元素对应 Excel 中的一行（forward-fill 后）

    处理逻辑:
      1. 根据 column_map 找到各字段对应的列 index
      2. 跳过表头行（第 1 行），从第 2 行开始读取数据
      3. 对 pd_name、target_file、iso_enable 三列做 forward-fill:
         如果某行的值为 None，则继承上一行的值
      4. signal_name 和 iso_value 为空的行跳过
    """
    from .config import reverse_map

    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 读取表头，建立 Excel列名 → 列index 的映射
    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=0):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    # 反转 column_map: {内部字段名: Excel列名} → {Excel列名: 内部字段名}
    if column_map is None:
        from .config import DEFAULT_COLUMN_MAP
        column_map = DEFAULT_COLUMN_MAP
    field_of_excel_col = {excel_name.strip(): field_name
                          for field_name, excel_name in column_map.items()}

    # 确定每个内部字段对应的列 index
    def get_col_idx(field_name: str) -> int:
        """根据内部字段名找到 Excel 列 index"""
        excel_col_name = column_map.get(field_name, "")
        return headers.get(excel_col_name, -1)

    idx_pd_name = get_col_idx("pd_name")
    idx_target_file = get_col_idx("target_file")
    idx_iso_enable = get_col_idx("iso_enable")
    idx_signal_name = get_col_idx("signal_name")
    idx_iso_value = get_col_idx("iso_value")

    # 验证必需的列都存在
    missing = []
    for field, idx in [("pd_name", idx_pd_name), ("signal_name", idx_signal_name),
                        ("iso_value", idx_iso_value), ("iso_enable", idx_iso_enable),
                        ("target_file", idx_target_file)]:
        if idx < 0:
            missing.append(f"{field} ({column_map.get(field, '?')})")
    if missing:
        raise ValueError(f"Excel 中找不到以下列: {', '.join(missing)}。"
                         f"表头: {list(headers.keys())}")

    # 逐行读取数据，做 forward-fill
    signals: List[IsoSignal] = []
    last_pd_name = None
    last_target_file = None
    last_iso_enable = None

    for row_idx in range(2, ws.max_row + 1):
        def cell_val(col_idx: int):
            """安全读取单元格值"""
            if col_idx < 0:
                return None
            return ws.cell(row=row_idx, column=col_idx + 1).value

        # 读取各列原始值
        raw_pd_name = cell_val(idx_pd_name)
        raw_target_file = cell_val(idx_target_file)
        raw_iso_enable = cell_val(idx_iso_enable)
        raw_signal_name = cell_val(idx_signal_name)
        raw_iso_value = cell_val(idx_iso_value)

        # Forward-fill: 空值（None 或空字符串）继承上一行
        # openpyxl 空单元格可能返回 None 或 ''，两种都要处理
        if raw_pd_name is not None and str(raw_pd_name).strip() != "":
            last_pd_name = str(raw_pd_name).strip()
        if raw_target_file is not None and str(raw_target_file).strip() != "":
            last_target_file = str(raw_target_file).strip()
        if raw_iso_enable is not None and str(raw_iso_enable).strip() != "":
            last_iso_enable = str(raw_iso_enable).strip()

        # signal_name 为空的行跳过（如空行）
        if raw_signal_name is None:
            continue

        signal_name = str(raw_signal_name).strip()
        if not signal_name:
            continue

        # 解析 iso_value: 支持数字和字符串
        iso_val = raw_iso_value
        if isinstance(iso_val, str):
            iso_val = iso_val.strip()
            try:
                iso_val = int(iso_val)
            except ValueError:
                continue  # 非数字的 tie 值行跳过
        if iso_val is None:
            continue
        iso_val = int(iso_val)
        if iso_val not in (0, 1):
            continue

        signals.append(IsoSignal(
            signal_name=signal_name,
            iso_enable=last_iso_enable or "",
            iso_value=iso_val,
            pd_name=last_pd_name or "",
            target_file=last_target_file or "",
        ))

    wb.close()
    return signals
